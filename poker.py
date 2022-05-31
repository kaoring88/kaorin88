from flask import (Flask, render_template, request, session)
import random
import openpyxl
import os

app=Flask(__name__)

key=os.urandom(21)
app.secret_key=key


H2=["heart",2]
S2=["spead",2]
D2=["daiya",2]
C2=["clover",2]

H3=["heart",3]
S3=["spead",3]
D3=["daiya",3]
C3=["clover",3]

H4=["heart",4]
S4=["spead",4]
D4=["daiya",4]
C4=["clover",4]

H5=["heart",5]
S5=["spead",5]
D5=["daiya",5]
C5=["clover",5]

H6=["heart",6]
S6=["spead",6]
D6=["daiya",6]
C6=["clover",6]

H7=["heart",7]
S7=["spead",7]
D7=["daiya",7]
C7=["clover",7]

H8=["heart",8]
S8=["spead",8]
D8=["daiya",8]
C8=["clover",8]

H9=["heart",9]
S9=["spead",9]
D9=["daiya",9]
C9=["clover",9]

H10=["heart",10]
S10=["spead",10]
D10=["daiya",10]
C10=["clover",10]

H11=["heart",11]
S11=["spead",11]
D11=["daiya",11]
C11=["clover",11]

H12=["heart",12]
S12=["spead",12]
D12=["daiya",12]
C12=["clover",12]

H13=["heart",13]
S13=["spead",13]
D13=["daiya",13]
C13=["clover",13]

HA14=["heart",14]
SA14=["spead",14]
DA14=["daiya",14]
CA14=["clover",14]

strdict={"H2":H2,"S2":S2,"D2":D2,"C2":C2,"H3":H3,"S3":S3,"D3":D3,"C3":C3,"H4":H4,"S4":S4,"D4":D4,"C4":C4,"H5":H5,"S5":S5,"D5":D5,"C5":C5,"H6":H6,"S6":S6,"D6":D6,"C6":C6,"H7":H7,"S7":S7,"D7":D7,"C7":C7,"H8":H8,"S8":S8,"D8":D8,"C8":C8,"H9":H9,"S9":S9,"D9":D9,"C9":C9,"H10":H10,"S10":S10,"D10":D10,"C10":C10,"H11":H11,"S11":S11,"D11":D11,"C11":C11,"H12":H12,"S12":S12,"D12":D12,"C12":C12,"H13":H13,"S13":S13,"D13":D13,"C13":C13,"HA14":HA14,"SA14":SA14,"DA14":DA14,"CA14":CA14}

@app.route("/")
def open():
    List=list(range(1,1000000,3))
    L=random.choice(List)
    My=[1,2,3,4,5]
    wb=openpyxl.load_workbook("card.xlsx")
    ws=wb["card"]
    s1="A{0}".format(L+1)
    ws[s1]=0
    s2="B{0}".format(L+1)
    ws[s2]=0
    for i in range(60):
        ws.cell(L+2,i+1,value=0)
    score=0
    scoreC=0
    a1="A{0}".format(L)
    b1="B{0}".format(L)
    c1="C{0}".format(L)
    d1="D{0}".format(L)
    e1="E{0}".format(L)
    for i in range(60):
        ws.cell(L+2,i+1,value=0)
    ws[a1]=None
    ws[b1]=None
    ws[c1]=None
    ws[d1]=None
    ws[e1]=None
    wb.save("card.xlsx")
    session["L"]=L
    session["My"]=My
    return render_template("open.html")

@app.route("/start")
def start():
    L=session.get("L")
    My=session.get("My")
    wb=openpyxl.load_workbook("card.xlsx")
    ws=wb["card"]
    s=wb.worksheets[0]
    a3="A{0}".format(L+2)
    if ws[a3].value==0 or ws[a3].value==None:
        xl1=list(strdict.keys())
        xl2=random.sample(xl1, len(xl1))
        for i in range(60):
            ws.cell(L+2,i+1,value=0)
        for i in range(0,len(xl2)):
            ws.cell(L+2,i+1,value=xl2[i])
        wb.save("card.xlsx")
    y=[]
    a3="A{0}".format(L+2)
    bz3="BZ{0}".format(L+2)
    abz="{0}:{1}".format(a3,bz3)
    for row in s[abz]:
        for col in row:
            if col.value==0 or col.value==None:
                pass
            else:
                y.append(col.value)
    M=[]
    for i in range(5):
        m=y.pop()
        M.append(m)
        My[i]=strdict[m]
    My00=M[0]
    My11=M[1]
    My22=M[2]
    My33=M[3]
    My44=M[4]
    a1="A{0}".format(L)
    b1="B{0}".format(L)
    c1="C{0}".format(L)
    d1="D{0}".format(L)
    e1="E{0}".format(L)
    for i in range(60):
        ws.cell(L+2,i+1,value=0)
    for i in range(0,len(y)):
        ws.cell(L+2,i+1,value=y[i])
    f1="F{0}".format(L)
    g1="G{0}".format(L)
    h1="H{0}".format(L)
    i1="I{0}".format(L)
    j1="J{0}".format(L)
    ws[f1]=My00
    ws[g1]=My11
    ws[h1]=My22
    ws[i1]=My33
    ws[j1]=My44
    ws[a1]=None
    ws[b1]=None
    ws[c1]=None
    ws[d1]=None
    ws[e1]=None
    wb.save("card.xlsx")
    a=len(y)
    return render_template("start.html",My00=My00,My11=My11,My22=My22,My33=My33,My44=My44,a=a,y=y,m=m,My=My,My0=My[0],My1=My[1],My2=My[2],My3=My[3],My4=My[4],H2=H2,S2=S2,D2=D2,C2=C2,H3=H3,S3=S3,D3=D3,C3=C3,H4=H4,S4=S4,D4=D4,C4=C4,H5=H5,S5=S5,D5=D5,C5=C5,H6=H6,S6=S6,D6=D6,C6=C6,H7=H7,S7=S7,D7=D7,C7=C7,H8=H8,S8=S8,D8=D8,C8=C8,H9=H9,S9=S9,D9=D9,C9=C9,H10=H10,S10=S10,D10=D10,C10=C10,H11=H11,S11=S11,D11=D11,C11=C11,H12=H12,S12=S12,D12=D12,C12=C12,H13=H13,S13=S13,D13=D13,C13=C13,HA14=HA14,SA14=SA14,DA14=DA14,CA14=CA14)

@app.route("/start1", methods=["POST"])
def change():
    L=session.get("L")
    My=session.get("My")
    wb=openpyxl.load_workbook("card.xlsx")
    ws=wb["card"]
    s=wb.worksheets[0]
    a1="A{0}".format(L)
    b1="B{0}".format(L)
    c1="C{0}".format(L)
    d1="D{0}".format(L)
    e1="E{0}".format(L)
    f1="F{0}".format(L)
    g1="G{0}".format(L)
    h1="H{0}".format(L)
    i1="I{0}".format(L)
    j1="J{0}".format(L)
    My[0]=strdict[ws[f1].value]
    My[1]=strdict[ws[g1].value]
    My[2]=strdict[ws[h1].value]
    My[3]=strdict[ws[i1].value]
    My[4]=strdict[ws[j1].value]
    My00=None
    My11=None
    My22=None
    My33=None
    My44=None
    My00=request.form.get("My00")
    My11=request.form.get("My11")
    My22=request.form.get("My22")
    My33=request.form.get("My33")
    My44=request.form.get("My44")
    y=[]
    a3="A{0}".format(L+2)
    bz3="BZ{0}".format(L+2)
    abz="{0}:{1}".format(a3,bz3)
    for row in s[abz]:
        for col in row:
            if col.value==0 or col.value==None:
                pass
            else:
                y.append(col.value)
    if My00!=None:
        ws[a1]=My00
    if My11!=None:
        ws[b1]=My11
    if My22!=None:
        ws[c1]=My22
    if My33!=None:
        ws[d1]=My33
    if My44!=None:
        ws[e1]=My44
    My00=ws[a1].value
    My11=ws[b1].value
    My22=ws[c1].value
    My33=ws[d1].value
    My44=ws[e1].value
    wb.save("card.xlsx")
    a=len(y)
    return render_template("start1.html",a=a,y=y,My00=My00,My11=My11,My22=My22,My33=My33,My44=My44,My0=My[0],My1=My[1],My2=My[2],My3=My[3],My4=My[4],H2=H2,S2=S2,D2=D2,C2=C2,H3=H3,S3=S3,D3=D3,C3=C3,H4=H4,S4=S4,D4=D4,C4=C4,H5=H5,S5=S5,D5=D5,C5=C5,H6=H6,S6=S6,D6=D6,C6=C6,H7=H7,S7=S7,D7=D7,C7=C7,H8=H8,S8=S8,D8=D8,C8=C8,H9=H9,S9=S9,D9=D9,C9=C9,H10=H10,S10=S10,D10=D10,C10=C10,H11=H11,S11=S11,D11=D11,C11=C11,H12=H12,S12=S12,D12=D12,C12=C12,H13=H13,S13=S13,D13=D13,C13=C13,HA14=HA14,SA14=SA14,DA14=DA14,CA14=CA14)

@app.route("/result", methods=["POST"])
def result():
    L=session.get("L")
    My=session.get("My")
    wb=openpyxl.load_workbook("card.xlsx")
    ws=wb["card"]
    s=wb.worksheets[0]
    f1="F{0}".format(L)
    g1="G{0}".format(L)
    h1="H{0}".format(L)
    i1="I{0}".format(L)
    j1="J{0}".format(L)
    My[0]=strdict[ws[f1].value]
    My[1]=strdict[ws[g1].value]
    My[2]=strdict[ws[h1].value]
    My[3]=strdict[ws[i1].value]
    My[4]=strdict[ws[j1].value]
    y=[]
    a3="A{0}".format(L+2)
    bz3="BZ{0}".format(L+2)
    abz="{0}:{1}".format(a3,bz3)
    for row in s[abz]:
        for col in row:
            if col.value==0 or col.value==None:
                pass
            else:
                y.append(col.value)
    a1="A{0}".format(L)
    b1="B{0}".format(L)
    c1="C{0}".format(L)
    d1="D{0}".format(L)
    e1="E{0}".format(L)    
    if ws[a1].value!=None:
        My000=y.pop()
        My[0]=strdict[My000]
        ws[a1]=None
    if ws[b1].value!=None:
        My111=y.pop()
        My[1]=strdict[My111]
        ws[b1]=None
    if ws[c1].value!=None:
        My222=y.pop()
        My[2]=strdict[My222]
        ws[c1]=None
    if ws[d1].value!=None:
        My333=y.pop()
        My[3]=strdict[My333]
        ws[d1]=None
    if ws[e1].value!=None:
        My444=y.pop()
        My[4]=strdict[My444]
        ws[e1]=None 
    for i in range(60):
        ws.cell(L+2,i+1,value=0)
    for i in range(0,len(y)):
        ws.cell(L+2,i+1,value=y[i])      
    result=request.form.get("result")
    syuruiM=[]
    suujiM=[]
    fourcardM=0
    fullhouseM=0
    threecardM=0
    twopairM=0
    onepairM=0
    nopairM=0
    for m in My:
        syuruiM.append(m[0])
        suujiM.append(m[1])
    for sm1 in suujiM:
        for sm2 in suujiM:
            if sm1==sm2:
                fourcardM+=1
            else:
                pass
    for sm3 in suujiM:
        for sm4 in suujiM:
            if sm3==sm4:
                fullhouseM+=1
            else:
                pass
    for sm5 in suujiM:
        for sm6 in suujiM:
            if sm5==sm6:
                threecardM+=1
            else:
                pass
    for sm7 in suujiM:
        for sm8 in suujiM:
            if sm7==sm8:
                twopairM+=1
            else:
                pass
    for sm9 in suujiM:
        for sm10 in suujiM:
            if sm9==sm10:
                onepairM+=1
            else:
                pass
    for sm11 in suujiM:
        for sm12 in suujiM:
            if sm11==sm12:
                nopairM+=1
            else:
                pass
    suujiMS=sorted(suujiM)
    if syuruiM[0]==syuruiM[1] and syuruiM[0]==syuruiM[2] and syuruiM[0]==syuruiM[3] and syuruiM[0]==syuruiM[4]:#種類が同じか
        if suujiMS[0]==10 and suujiMS[1]==11 and suujiMS[2]==12 and suujiMS[3]==13 and suujiMS[4]==14:#ロイヤルか
            sc=1000
        elif suujiMS[0]+1==suujiMS[1] and suujiMS[0]+2==suujiMS[2] and suujiMS[0]+3==suujiMS[3] and suujiMS[0]+4==suujiMS[4]:#ストレートか
            sc=350
        elif fourcardM==17:
            sc=200
        elif fullhouseM==13:
            sc=100
        else:#フラッシュ
            sc=70

    if fourcardM==17:#フォアカード
        sc=200
    elif fullhouseM==13:#フルハウス
        sc=100
    elif suujiMS[0]+1==suujiMS[1] and suujiMS[0]+2==suujiMS[2] and suujiMS[0]+3==suujiMS[3] and suujiMS[0]+4==suujiMS[4]:
        sc=60
    elif threecardM==11:
        sc=50
    elif twopairM==9:
        sc=30
    elif onepairM==7:
        sc=10
    else:
        sc=0
    a2="A{0}".format(L+1)
    score=ws[a2].value+sc
    ws[a2]=score
    dice=["ロイヤル","ストフラ","フォア","フル","フラッシュ","ストレート","スリー","ツー","ワン","ノー"]
    w=[154,1390,24000,14410,19650,39250,211280,475390,4225690,5011770]
    Cpu=random.choices(dice, k=1, weights=w)
    if Cpu[0]=="ロイヤル":
        CpuT=1000
    elif Cpu[0]=="ストフラ":
        CpuT=350
    elif Cpu[0]=="フォア":
        CpuT=200
    elif Cpu[0]=="フル":
        CpuT=100
    elif Cpu[0]=="フラッシュ":
        CpuT=70
    elif Cpu[0]=="ストレート":
        CpuT=60
    elif Cpu[0]=="スリー":
        CpuT=50
    elif Cpu[0]=="ツー":
        CpuT=30
    elif Cpu[0]=='ワン':
        CpuT=10
    else:
        CpuT=0
    b2="B{0}".format(L+1)
    scoreC=ws[b2].value+CpuT
    ws[b2]=scoreC
    wb.save("card.xlsx")
    chosei=[10,11,12,13,14,15,16,17,18,19,20]
    if len(y)>=random.choice(chosei):
        a=len(y)
        return render_template("result.html",CpuT=CpuT,sc=sc,scoreC=ws[b2].value,score=ws[a2].value,y=y,a=a,suujiMS=suujiMS,result=result,My=My,fourcardM=fourcardM,fullhouseM=fullhouseM,threecardM=threecardM,twopairM=twopairM,onepairM=onepairM,nopairM=nopairM,syuruiM=syuruiM,suujiM=suujiM,My0=My[0],My1=My[1],My2=My[2],My3=My[3],My4=My[4],H2=H2,S2=S2,D2=D2,C2=C2,H3=H3,S3=S3,D3=D3,C3=C3,H4=H4,S4=S4,D4=D4,C4=C4,H5=H5,S5=S5,D5=D5,C5=C5,H6=H6,S6=S6,D6=D6,C6=C6,H7=H7,S7=S7,D7=D7,C7=C7,H8=H8,S8=S8,D8=D8,C8=C8,H9=H9,S9=S9,D9=D9,C9=C9,H10=H10,S10=S10,D10=D10,C10=C10,H11=H11,S11=S11,D11=D11,C11=C11,H12=H12,S12=S12,D12=D12,C12=C12,H13=H13,S13=S13,D13=D13,C13=C13,HA14=HA14,SA14=SA14,DA14=DA14,CA14=CA14)
    else:
        return render_template("result2.html",CpuT=CpuT,sc=sc,scoreC=ws[b2].value,score=ws[a2].value,y=y,suujiMS=suujiMS,result=result,My=My,fourcardM=fourcardM,fullhouseM=fullhouseM,threecardM=threecardM,twopairM=twopairM,onepairM=onepairM,nopairM=nopairM,syuruiM=syuruiM,suujiM=suujiM,My0=My[0],My1=My[1],My2=My[2],My3=My[3],My4=My[4],H2=H2,S2=S2,D2=D2,C2=C2,H3=H3,S3=S3,D3=D3,C3=C3,H4=H4,S4=S4,D4=D4,C4=C4,H5=H5,S5=S5,D5=D5,C5=C5,H6=H6,S6=S6,D6=D6,C6=C6,H7=H7,S7=S7,D7=D7,C7=C7,H8=H8,S8=S8,D8=D8,C8=C8,H9=H9,S9=S9,D9=D9,C9=C9,H10=H10,S10=S10,D10=D10,C10=C10,H11=H11,S11=S11,D11=D11,C11=C11,H12=H12,S12=S12,D12=D12,C12=C12,H13=H13,S13=S13,D13=D13,C13=C13,HA14=HA14,SA14=SA14,DA14=DA14,CA14=CA14)

@app.route("/miinya", methods=["GET"])
def miinya():
    L=session.get("L")
    wb=openpyxl.load_workbook("card.xlsx")
    ws=wb["card"]
    for i in range(60):
        ws.cell(L+2,i+1,value=0)
    a2="A{0}".format(L+1)
    b2="B{0}".format(L+1)
    ws[a2]=0
    ws[b2]=0
    wb.save("card.xlsx")
    return render_template("miinya.html")

if __name__=="__main__":
    app.run(host="0.0.0.0", port=5000)
    